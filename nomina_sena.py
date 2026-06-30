import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Nómina Junio 2026"

empleados = [
    ["1012345678", "Carlos Martínez", "Desarrollador", 2500000, 15],
    ["1023456789", "Laura Gómez", "Analista QA", 2800000, 10],
    ["1034567890", "Juan Rodríguez", "Diseñador", 2200000, 20],
    ["1045678901", "Ana López", "Scrum Master", 3500000, 5],
    ["1056789012", "Pedro García", "DevOps", 3200000, 12],
    ["1067890123", "Diana Torres", "FullStack", 2900000, 18],
    ["1078901234", "Luis Ramírez", "Analista Datos", 2600000, 8],
    ["1089012345", "María Herrera", "Backend", 3100000, 14],
    ["1090123456", "Andrés Castro", "Frontend", 2400000, 22],
    ["1101234567", "Valentina Morales", "Project Manager", 3800000, 0],
]

encabezados = [
    "Documento", "Nombre", "Cargo", "Salario", "Horas Extras",
    "Auxilio Transporte", "Valor Hora", "Valor Hora Extra",
    "Total Horas Extras", "Devengado", "Salud", "Pensión", "Neto a Pagar"
]

for col, titulo in enumerate(encabezados, 1):
    ws.cell(row=1, column=col, value=titulo)

for i, emp in enumerate(empleados, 1):
    fila = i + 1
    doc, nombre, cargo, salario, h_ext = emp
    auxilio = 200000 if salario <= 2847000 else 0
    v_hora = salario / 240
    v_hora_ext = v_hora * 1.25
    total_h_ext = h_ext * v_hora_ext
    devengado = salario + auxilio + total_h_ext
    salud = devengado * 0.04
    pension = devengado * 0.04
    neto = devengado - salud - pension

    ws.cell(row=fila, column=1, value=doc)
    ws.cell(row=fila, column=2, value=nombre)
    ws.cell(row=fila, column=3, value=cargo)
    ws.cell(row=fila, column=4, value=salario)
    ws.cell(row=fila, column=5, value=h_ext)
    ws.cell(row=fila, column=6, value=auxilio)
    ws.cell(row=fila, column=7, value=v_hora)
    ws.cell(row=fila, column=8, value=v_hora_ext)
    ws.cell(row=fila, column=9, value=total_h_ext)
    ws.cell(row=fila, column=10, value=devengado)
    ws.cell(row=fila, column=11, value=salud)
    ws.cell(row=fila, column=12, value=pension)
    ws.cell(row=fila, column=13, value=neto)

# Estilos
azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
blanco = Font(color="FFFFFF", bold=True)
negro = Font(bold=True)
borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
pesos = '"$"#,##0.00'

filas = len(empleados) + 1
columnas = len(encabezados)

for r in range(1, filas + 1):
    for c in range(1, columnas + 1):
        celda = ws.cell(row=r, column=c)
        celda.border = borde
        if r == 1:
            celda.font = blanco
            celda.fill = azul
        else:
            celda.font = negro
        if c in [4, 6, 7, 8, 9, 10, 11, 12, 13] and r > 1:
            celda.number_format = pesos

# Ajuste automático de columnas
for c in range(1, columnas + 1):
    max_len = 0
    letra = get_column_letter(c)
    for r in range(1, filas + 1):
        val = str(ws.cell(row=r, column=c).value or "")
        if len(val) > max_len:
            max_len = len(val)
    ancho = max_len + 3
    if c in [4, 6, 7, 8, 9, 10, 11, 12, 13]:
        ancho = max(ancho, 18)
    ws.column_dimensions[letra].width = ancho

# Hoja Resumen Gerencial
ws2 = wb.create_sheet("Resumen Gerencial")
ws2["A1"] = "RESUMEN GERENCIAL"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws2.merge_cells("A1:B1")

ws2["A3"] = "INDICADOR"
ws2["B3"] = "VALOR"
for celda in [ws2["A3"], ws2["B3"]]:
    celda.font = blanco
    celda.fill = azul
    celda.border = borde

hoja_nomina = "'Nómina Junio 2026'"
indicadores = [
    ("Total Nómina", f"=SUM({hoja_nomina}!M2:M11)"),
    ("Promedio Salarial", f"=AVERAGE({hoja_nomina}!D2:D11)"),
    ("Mayor Salario", f"=MAX({hoja_nomina}!D2:D11)"),
    ("Menor Salario", f"=MIN({hoja_nomina}!D2:D11)"),
]

for i, (ind, formula) in enumerate(indicadores, 4):
    ws2.cell(row=i, column=1, value=ind).border = borde
    ws2.cell(row=i, column=2, value=formula).border = borde
    ws2.cell(row=i, column=2).number_format = pesos
    ws2.cell(row=i, column=1).font = negro

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 28

wb.save("nomina_empresa.xlsx")
print("Archivo creado: nomina_empresa.xlsx")

# Cálculo de respuestas al análisis
total_nomina = 0
total_salud = 0
total_pension = 0
mayor = 0
empleado_top = ""

for emp in empleados:
    salario, h_ext = emp[3], emp[4]
    auxilio = 200000 if salario <= 2847000 else 0
    v_hora = salario / 240
    v_hora_ext = v_hora * 1.25
    total_h_ext = h_ext * v_hora_ext
    devengado = salario + auxilio + total_h_ext
    salud_val = devengado * 0.04
    pension_val = devengado * 0.04
    neto = devengado - salud_val - pension_val
    total_nomina += neto
    total_salud += salud_val
    total_pension += pension_val
    if neto > mayor:
        mayor = neto
        empleado_top = emp[1]

print(f"1. Total pagado en nómina: ${total_nomina:,.2f}")
print(f"2. Total descontado por salud: ${total_salud:,.2f}")
print(f"3. Total descontado por pensión: ${total_pension:,.2f}")
print(f"4. Empleado con mayor pago: {empleado_top} (${mayor:,.2f})")
print("5. Ventajas de automatizar la nómina mediante Python:")
print("   - Reduce errores humanos propios del cálculo manual.")
print("   - Ahorra tiempo al generar la nómina y los reportes en segundos.")
print("   - Es reproducible: se puede reutilizar cada mes con los mismos empleados o datos nuevos.")
print("   - Permite generar archivos Excel con formato profesional automáticamente.")
print("   - Facilita el análisis y la toma de decisiones mediante resúmenes gerenciales.")
